import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
from tkinter.filedialog import askopenfile
from tkinter import *
from openpyxl import load_workbook

root = Tk()
root.geometry("1530x830")
root.title("Email Sender")
root.configure(bg="lightblue", padx=10, pady=10)


sender_address = StringVar()
sender_pass = StringVar()
fileAttached = StringVar()
subject = StringVar()
content = StringVar()
totalmail = StringVar()
sentlst = StringVar()

sender_address.set("youremail@gmail.com")
sender_pass.set("yourpassword")
fileAttached.set("No file selected")
subject.set("Testing is going on....")
content.set("I am testing my application")
totalmail.set("Select file to get mails")
sentlst.set("Empty list")

subject_text = subject.get()

mail_lst = []
def open_file():
    inp_file = askopenfile(mode ='r')
    workbook = load_workbook(inp_file.name)
    current = workbook.active
    no_rows = current.max_row
    data = ""
    for i in range(1, no_rows + 1):
        cell_content = current.cell(row=i, column=1)
        # print(cell_content.value)
        mail_lst.append(cell_content.value)
        data += f" {i}. {cell_content.value} \n"
    totalmail.set(data)
    mailList.delete("1.0", "end")
    mailList.insert(END, totalmail.get())


class SendMail():
    def __init__(self, sender, passw, lst):
        self.sender_address = sender
        self.sender_pass = passw
        self.lst = lst
        self.message = MIMEMultipart()


    def selectFile(self):
        inpfile = askopenfile("r")
        with open(inpfile.name, "rb") as attachment:
            part = MIMEBase("application", "octet-stream")
            part.set_payload(attachment.read())

        encoders.encode_base64(part)

        part.add_header('Content-Disposition', "attachment; filename= %s" % inpfile.name)

        self.message.attach(part)
        fileAttached.set(inpfile.name)

    def sendmail(self):
        l = self.lst
        sentlist = ""
        for mail in l:
            receiver = mail
            self.message['From'] = self.sender_address

            self.message['Subject'] = subjectInput.get(1.0, "end-1c")

            self.mail_content = contentInput.get(1.0, "end-1c")

            self.message.attach(MIMEText(self.mail_content, 'plain'))
            s = smtplib.SMTP('smtp.gmail.com', 587)
            s.starttls()
            s.login(self.sender_address, self.sender_pass)
            text = self.message.as_string()
            s.sendmail(self.sender_address,receiver, text)
            print(f"Mail sent to {receiver}")
            sentlist += f"Mail sent to {receiver} \n"
            # sent_mailList.insert(END, sentlst.get())
            s.quit()
        sentlst.set(sentlist)
        sent_mailList.delete("1.0", "end")
        sent_mailList.insert(END, sentlst.get())
        








headerFrame = Frame(root, bg="lightblue", pady=10, padx=20)
headerFrame.pack(fill=X)

headerLabel = Label(headerFrame, bg="lightblue", text="Select .excel file of mails", pady=10, padx=20,
                    font=("Arial", 15))
headerLabel.pack(side=LEFT, padx=20)

headerButton = Button(headerFrame, text="Select File", pady=8, padx=20, command=lambda:open_file())
headerButton.pack(side=LEFT)



mailListFrame = Frame(root, bg="Pale Green1",width=715, height=80)
mailListFrame.pack(side=LEFT, fill=BOTH, padx=20)

totalMailLabel = Label(mailListFrame, text="Total Mails : ")
totalMailLabel.pack(side=TOP, anchor=W)

mailList = Text(mailListFrame)
mailList.pack(side=TOP, fill=X, pady=10)
mailList.insert(END, totalmail.get())

sentmailLabel = Label(mailListFrame, text="Sent Mail List : ")
sentmailLabel.pack(side=TOP, anchor=W)

sent_mailList = Text(mailListFrame)
sent_mailList.pack(side=TOP, fill=X)
sent_mailList.insert(END, sentlst.get())




mailDraftFrame = Frame(root, bg="Pale Green1", width=715, height=80)
mailDraftFrame.pack(fill=BOTH, side=LEFT)


f1 = Frame(mailDraftFrame)
f1.pack(fill=X, side=TOP)

senderEmailAddress = Label(f1, text="Sender's mail address : ", width=400, height=2)
senderEmailAddress.pack(side=TOP, fill=X)

senderEmailAddressInput = Entry(f1, font=("Arial"), textvariable=sender_address)
senderEmailAddressInput.pack(side=TOP, fill=X)

# f6 = Frame(mailDraftFrame)
# f6.pack(fill=X, side=TOP)

# senderEmailPassword = Label(f6, text="Sender's mail password : ", width=400, height=2)
# senderEmailPassword.pack(side=TOP, fill=X)

# senderEmailPasswordInput = Entry(f6, font=("Arial"), textvariable=sender_pass, show="*")
# senderEmailPasswordInput.pack(side=TOP, fill=X)



f2 = Frame(mailDraftFrame)
f2.pack(fill=X, side=TOP, pady=10)

subjectLabel = Label(f2, text="Enter Subject : ", width=50, height=2)
subjectLabel.pack(side=TOP, fill=X)

subjectInput = Text(f2, height=5)
subjectInput.pack(side=TOP, fill=X)
subjectInput.insert(END, subject.get())



f3 = Frame(mailDraftFrame)
f3.pack(side=TOP, fill=X, pady=10)

contentLabel = Label(f3, text="Content")
contentLabel.pack(side=TOP, fill=X)

contentInput = Text(f3, height=20)
contentInput.pack(side=TOP, fill=X)
contentInput.insert(END, content.get())



f4 = Frame(mailDraftFrame)
f4.pack(side=TOP, fill=X, pady=10)

inputfileInput = Entry(f4, textvariable=fileAttached)
inputfileInput.pack(side=TOP, fill=X, ipady=5)

start = SendMail(sender_address.get(), sender_pass.get(), mail_lst)

fileSelectButton = Button(f4, text="Select Attachment", command=start.selectFile)
fileSelectButton.pack(side=TOP, fill=X)

f5 = Frame(mailDraftFrame, bg="Pale Green1")
f5.pack(side=TOP, fill=X)

sendButton = Button(f5, text="Send", command=start.sendmail, width=10, height=4, font=("Arial", 12))
sendButton.pack(side=TOP, pady=10)



root.mainloop()
